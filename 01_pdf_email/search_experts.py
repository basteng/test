#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
自动搜索专家信息
从Email地址搜索专家的姓名、机构、院系、研究领域等信息
"""

import os
import sys
import json
import time
from openpyxl import load_workbook
from datetime import datetime

# 注意：此脚本需要手动使用WebSearch，这里提供搜索逻辑
# 实际搜索需要用户手动执行或使用API

def extract_domain(email):
    """从email地址提取域名"""
    if '@' in email:
        return email.split('@')[1].lower()
    return "unknown"

def extract_name_from_email(email):
    """从email地址推断可能的姓名"""
    if '@' not in email:
        return ""

    username = email.split('@')[0]
    # 处理常见格式：john.doe, j.doe, jdoe等
    parts = username.replace('.', ' ').replace('-', ' ').replace('_', ' ').split()
    return ' '.join([p.capitalize() for p in parts if len(p) > 1])

def create_expert_template(input_file):
    """创建专家信息模板JSON"""

    if not os.path.exists(input_file):
        print(f"错误：文件不存在 - {input_file}")
        return False

    print("="*80)
    print("专家信息搜索工具")
    print("="*80)
    print(f"\n正在读取文件: {input_file}")

    # 加载Excel文件
    try:
        wb = load_workbook(input_file, data_only=True)
    except Exception as e:
        print(f"错误：无法读取Excel文件 - {str(e)}")
        return False

    # 检查是否存在"去重Email合并视图"工作表
    if "去重Email合并视图" not in wb.sheetnames:
        print("错误：未找到'去重Email合并视图'工作表")
        return False

    ws = wb["去重Email合并视图"]

    # 读取email列表
    print("正在读取Email列表...")
    expert_info = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:
            continue

        serial = row[0]
        email = row[1]
        count = row[2]
        first_file = row[3]

        domain = extract_domain(email)
        suggested_name = extract_name_from_email(email)

        # 创建模板
        expert_info[email] = {
            "email": email,
            "name": suggested_name if suggested_name else "待填写",
            "title": "待填写",
            "institution": f"推断自域名: {domain}",
            "institution_short": domain.split('.')[0].upper() if '.' in domain else domain,
            "department": "待填写",
            "department_short": "",
            "research_areas": [],
            "research_interests": "待填写",
            "h_index": None,
            "personal_website": "",
            "google_scholar": "",
            "verified": False,
            "confidence": "low",
            "search_queries": [
                f'"{email}" researcher',
                f'"{email}" professor',
                f'{suggested_name} {domain.split(".")[0]} university' if suggested_name else ""
            ],
            "serial_number": serial,
            "publication_count": count,
            "first_file": first_file,
            "last_updated": datetime.now().strftime('%Y-%m-%d')
        }

    print(f"已创建 {len(expert_info)} 个专家信息模板")

    # 保存JSON文件
    output_file = "expert_info_template.json"
    print(f"\n正在保存模板: {output_file}")

    try:
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(expert_info, f, ensure_ascii=False, indent=2)

        print("✓ 模板文件已生成！")
        print("="*80)
        print("\n使用说明:")
        print("1. 打开 expert_info_template.json")
        print("2. 对于每个email，可以：")
        print("   - 使用提供的search_queries搜索")
        print("   - 手动填写专家信息")
        print("   - 修改confidence为'high'表示确认")
        print("3. 保存文件后运行 generate_markdown_detailed.py 生成Markdown")
        print()
        print("提示：你也可以使用我已搜索的示例结果作为参考")

        return True

    except Exception as e:
        print(f"错误：无法保存JSON文件 - {str(e)}")
        return False

def main():
    """主函数"""

    # 确定输入文件
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
    else:
        input_file = "PDF_Email提取报告_已过滤.xlsx"
        if not os.path.exists(input_file):
            print("使用方法:")
            print("  python search_experts.py [Excel文件路径]")
            print()
            print(f"错误：当前目录未找到 {input_file}")
            return

    # 创建模板
    success = create_expert_template(input_file)

    if success:
        print("\n下一步：")
        print("运行: python generate_markdown_detailed.py")

if __name__ == "__main__":
    main()
