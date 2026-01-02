#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成Markdown文件 - 按机构分类专家
从Email域名提取机构信息，按机构组织专家列表
"""

import os
import sys
from openpyxl import load_workbook
from collections import defaultdict
from datetime import datetime

def extract_domain(email):
    """从email地址提取域名"""
    if '@' in email:
        return email.split('@')[1].lower()
    return "unknown"

def generate_markdown(input_file):
    """生成Markdown文件"""

    if not os.path.exists(input_file):
        print(f"错误：文件不存在 - {input_file}")
        return False

    print("="*80)
    print("Markdown文件生成工具 - 按机构分类专家")
    print("="*80)
    print(f"\n正在读取文件: {input_file}")

    # 加载Excel文件
    try:
        wb = load_workbook(input_file, data_only=True)
    except Exception as e:
        print(f"错误：无法读取Excel文件 - {str(e)}")
        return False

    # 检查是否存在"去重Email合并视图"工作表
    if "去重Email合并视图" not in wb.sheetnames:
        print("错误：未找到'去重Email合并视图'工作表")
        print("请先运行 filter_edu_emails.py 生成该工作表")
        return False

    ws = wb["去重Email合并视图"]

    # 读取数据
    print("正在读取Email数据...")

    # 按域名分组存储email信息
    domain_to_emails = defaultdict(list)

    # 读取数据（跳过标题行）
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:  # 检查Email地址列（第2列）
            continue

        serial_number = row[0]  # 对应序号（工作表1）
        email = row[1]  # Email地址
        count = row[2]  # 出现次数
        first_file = row[3]  # 首次出现文件名
        all_files = row[6]  # 所有相关文献文件名

        # 提取域名
        domain = extract_domain(email)

        # 存储信息
        domain_to_emails[domain].append({
            'email': email,
            'serial': serial_number,
            'count': count,
            'first_file': first_file,
            'all_files': all_files
        })

    print(f"找到 {len(domain_to_emails)} 个不同的机构/域名")
    print(f"总Email数: {sum(len(emails) for emails in domain_to_emails.values())}")

    # 生成Markdown内容
    print("\n正在生成Markdown文件...")

    markdown_lines = []

    # 添加文件头
    markdown_lines.append("# 专家机构分类")
    markdown_lines.append("")
    markdown_lines.append(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    markdown_lines.append("")
    markdown_lines.append(f"**统计信息:**")
    markdown_lines.append(f"- 机构/域名总数: {len(domain_to_emails)}")
    markdown_lines.append(f"- 专家总数: {sum(len(emails) for emails in domain_to_emails.values())}")
    markdown_lines.append("")
    markdown_lines.append("---")
    markdown_lines.append("")

    # 按域名排序
    sorted_domains = sorted(domain_to_emails.items(), key=lambda x: x[0])

    # 为每个域名生成一个章节
    for domain, emails in sorted_domains:
        # 一级标题：域名
        markdown_lines.append(f"# {domain}")
        markdown_lines.append("")
        markdown_lines.append(f"**专家数量:** {len(emails)}")
        markdown_lines.append("")

        # 为每个email生成一个小节
        for email_info in emails:
            # 二级标题：email地址
            markdown_lines.append(f"## {email_info['email']}")
            markdown_lines.append("")

            # 详细信息
            markdown_lines.append(f"- **出现次数:** {email_info['count']}")
            markdown_lines.append(f"- **首次出现:** {email_info['first_file']} (序号: {email_info['serial']})")
            markdown_lines.append(f"- **相关文献:**")

            # 解析文件名列表
            if email_info['all_files']:
                files = [f.strip() for f in str(email_info['all_files']).split(';')]
                for file in files:
                    markdown_lines.append(f"  - {file}")

            markdown_lines.append("")

        markdown_lines.append("---")
        markdown_lines.append("")

    # 保存Markdown文件
    output_file = input_file.replace('_已过滤.xlsx', '_专家机构分类.md')
    if output_file == input_file:
        output_file = input_file.replace('.xlsx', '_专家机构分类.md')

    print(f"\n正在保存文件: {output_file}")
    try:
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write('\n'.join(markdown_lines))

        print("✓ Markdown文件生成成功！")
        print("="*80)
        print(f"\n文件详情:")
        print(f"  输出文件: {output_file}")
        print(f"  机构/域名数: {len(domain_to_emails)}")
        print(f"  专家总数: {sum(len(emails) for emails in domain_to_emails.values())}")

        # 显示前几个域名
        print(f"\n包含的域名:")
        for i, domain in enumerate(sorted(domain_to_emails.keys())[:10], 1):
            count = len(domain_to_emails[domain])
            print(f"  {i}. {domain} ({count}个专家)")

        if len(domain_to_emails) > 10:
            print(f"  ... 还有 {len(domain_to_emails) - 10} 个域名")

        return True

    except Exception as e:
        print(f"错误：无法保存Markdown文件 - {str(e)}")
        return False

def main():
    """主函数"""
    print("="*80)
    print("Markdown生成工具")
    print("="*80)
    print()

    # 确定输入文件
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
    else:
        # 默认处理当前目录下的已过滤文件
        input_file = "PDF_Email提取报告_已过滤.xlsx"
        if not os.path.exists(input_file):
            print("使用方法:")
            print("  python generate_markdown.py [Excel文件路径]")
            print()
            print("示例:")
            print("  python generate_markdown.py")
            print("    -> 处理当前目录的 PDF_Email提取报告_已过滤.xlsx")
            print()
            print("  python generate_markdown.py \"D:\\报告\\PDF_Email提取报告_已过滤.xlsx\"")
            print("    -> 处理指定路径的文件")
            print()
            print(f"错误：当前目录未找到 {input_file}")
            print()
            print("提示: 请先运行 filter_edu_emails.py 生成已过滤的Excel文件")
            return

    # 生成Markdown
    success = generate_markdown(input_file)

    if success:
        print("\n✓ 处理完成！")
        print("\n说明:")
        print("  - Markdown文件已按机构/域名分类")
        print("  - 一级标题为域名（机构）")
        print("  - 二级标题为专家email")
        print("  - 包含每个专家的文献列表")
    else:
        print("\n✗ 处理失败，请检查错误信息")

if __name__ == "__main__":
    main()
