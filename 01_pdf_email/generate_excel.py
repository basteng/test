#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成Excel报告 - PDF Email提取结果
将路径转换为Windows格式
"""

import os
import re
import glob
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def get_file_size_mb(file_path):
    """获取文件大小（MB）"""
    size_bytes = os.path.getsize(file_path)
    size_mb = size_bytes / (1024 * 1024)
    return size_mb

def linux_to_windows_path(linux_path):
    """将Linux路径转换为Windows路径"""
    # /mnt/d/... -> D:\...
    # /mnt/c/... -> C:\...
    if linux_path.startswith('/mnt/'):
        # 提取盘符
        parts = linux_path.split('/')
        if len(parts) >= 3:
            drive = parts[2].upper()
            # 重新组合路径
            rest_path = '/'.join(parts[3:])
            windows_path = f"{drive}:\\" + rest_path.replace('/', '\\')
            return windows_path
    # 如果不是/mnt/格式，直接替换/为\
    return linux_path.replace('/', '\\')

def extract_emails_from_pdf(pdf_path):
    """从PDF文件中提取email地址"""
    emails = []
    try:
        import PyPDF2
        with open(pdf_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            text = ""
            max_pages = min(5, len(pdf_reader.pages))
            for page_num in range(max_pages):
                page = pdf_reader.pages[page_num]
                text += page.extract_text()

        # 使用正则表达式提取email
        email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
        found_emails = re.findall(email_pattern, text)

        # 去重
        seen = set()
        for email in found_emails:
            email_lower = email.lower()
            if email_lower not in seen:
                seen.add(email_lower)
                emails.append(email)

    except Exception as e:
        print(f"处理文件 {pdf_path} 时出错: {str(e)}")
        return []

    return emails

def main():
    """主函数"""
    current_dir = os.getcwd()
    print("开始处理PDF文件并生成Excel报告...")
    print("="*80)
    print(f"当前工作目录: {current_dir}")
    print("正在递归搜索所有子文件夹中的PDF文件...")
    print("="*80)

    # 查找所有PDF文件（递归遍历所有子文件夹）
    pdf_files = glob.glob("**/*.pdf", recursive=True)

    if not pdf_files:
        print("未找到PDF文件")
        return

    print(f"找到 {len(pdf_files)} 个PDF文件\n")

    # 存储结果
    all_results = []  # 包含所有文件
    MAX_SIZE_MB = 30

    # 处理每个PDF文件
    for pdf_file in sorted(pdf_files):
        file_path = os.path.abspath(pdf_file)
        file_name = os.path.basename(pdf_file)
        file_size_mb = get_file_size_mb(pdf_file)
        windows_path = linux_to_windows_path(file_path)

        # 获取相对路径和文件夹路径
        relative_path = pdf_file
        folder_path = os.path.dirname(pdf_file) if os.path.dirname(pdf_file) else "."
        windows_folder = linux_to_windows_path(os.path.abspath(folder_path))

        # 检查文件大小
        if file_size_mb > MAX_SIZE_MB:
            all_results.append({
                'name': file_name,
                'folder': folder_path,
                'windows_folder': windows_folder,
                'path': windows_path,
                'size_mb': file_size_mb,
                'emails': [],
                'status': '跳过',
                'reason': f'文件大小 {file_size_mb:.2f} MB 超过 {MAX_SIZE_MB} MB 限制'
            })
            print(f"⊗ 跳过 ({file_size_mb:.2f} MB): {relative_path}")
            continue

        # 提取email
        print(f"处理中 ({file_size_mb:.2f} MB): {relative_path}")
        emails = extract_emails_from_pdf(pdf_file)

        all_results.append({
            'name': file_name,
            'folder': folder_path,
            'windows_folder': windows_folder,
            'path': windows_path,
            'size_mb': file_size_mb,
            'emails': emails,
            'status': '已处理',
            'reason': ''
        })

    print("="*80)
    print("\n生成Excel文件...")

    # 创建Excel工作簿
    wb = Workbook()

    # 第一个工作表：所有PDF文件
    ws1 = wb.active
    ws1.title = "所有PDF文件"

    # 设置样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    skipped_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")  # 黄色背景
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 写入标题行
    headers = ["序号", "文件名", "所在文件夹（相对路径）", "文件夹完整路径（Windows）", "文件完整路径（Windows）", "文件大小(MB)", "处理状态", "Email地址", "Email数量", "备注"]
    ws1.append(headers)

    # 设置标题行样式
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 写入数据
    for idx, result in enumerate(all_results, 1):
        email_str = "; ".join(result['emails']) if result['emails'] else "未找到"
        email_count = len(result['emails'])

        ws1.append([
            idx,
            result['name'],
            result['folder'],
            result['windows_folder'],
            result['path'],
            round(result['size_mb'], 2),
            result['status'],
            email_str,
            email_count,
            result['reason']
        ])

        # 设置单元格样式
        row_num = idx + 1
        for cell in ws1[row_num]:
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            # 如果是跳过的文件，设置黄色背景
            if result['status'] == '跳过':
                cell.fill = skipped_fill

    # 调整列宽
    ws1.column_dimensions['A'].width = 8
    ws1.column_dimensions['B'].width = 50
    ws1.column_dimensions['C'].width = 25
    ws1.column_dimensions['D'].width = 70
    ws1.column_dimensions['E'].width = 70
    ws1.column_dimensions['F'].width = 15
    ws1.column_dimensions['G'].width = 12
    ws1.column_dimensions['H'].width = 40
    ws1.column_dimensions['I'].width = 12
    ws1.column_dimensions['J'].width = 35

    # 第二个工作表：汇总统计
    ws2 = wb.create_sheet(title="统计汇总")

    # 统计信息
    processed_files = [r for r in all_results if r['status'] == '已处理']
    skipped_files = [r for r in all_results if r['status'] == '跳过']
    total_emails = sum(len(r['emails']) for r in all_results)
    files_with_emails = sum(1 for r in all_results if r['emails'])
    files_without_emails = sum(1 for r in processed_files if not r['emails'])

    summary_data = [
        ["项目", "数值"],
        ["总文件数", len(pdf_files)],
        ["成功处理文件数", len(processed_files)],
        ["跳过文件数", len(skipped_files)],
        ["提取到的Email总数", total_emails],
        ["有Email的文件数", files_with_emails],
        ["无Email的文件数", files_without_emails],
        ["生成时间", datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
    ]

    for row in summary_data:
        ws2.append(row)

    # 设置标题行样式
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 设置数据行样式
    for row in ws2.iter_rows(min_row=2, max_row=len(summary_data)):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='center')

    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 30

    # 保存Excel文件
    excel_filename = "PDF_Email提取报告.xlsx"
    wb.save(excel_filename)

    print(f"\n✓ Excel文件已生成: {linux_to_windows_path(os.path.abspath(excel_filename))}")
    print("="*80)
    print("\n文件包含2个工作表:")
    print("  1. 所有PDF文件 - 包含所有文件及其文件夹路径（超过30MB的文件以黄色背景标注）")
    print("  2. 统计汇总 - 整体统计信息")
    print(f"\n摘要:")
    print(f"  总文件数: {len(pdf_files)}")
    print(f"  成功处理: {len(processed_files)}")
    print(f"  跳过文件: {len(skipped_files)}")
    print(f"  提取Email总数: {total_emails}")
    print("\n提示: 脚本已递归搜索所有子文件夹中的PDF文件")

if __name__ == "__main__":
    main()
