#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
过滤EDU邮箱 - 从Excel报告中移除包含edu的邮箱
读取已有的Excel文件，添加新工作表显示过滤后的结果
"""

import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def filter_edu_emails(email_list):
    """过滤掉包含edu的邮箱地址（不区分大小写）"""
    if not email_list:
        return []

    # email_list 可能是字符串（以分号分隔）或列表
    if isinstance(email_list, str):
        if email_list == "未找到":
            return []
        emails = [e.strip() for e in email_list.split(';')]
    else:
        emails = email_list

    # 过滤掉包含edu的邮箱
    filtered = [email for email in emails if 'edu' not in email.lower()]
    return filtered

def process_excel(input_file):
    """处理Excel文件，添加过滤后的工作表"""

    if not os.path.exists(input_file):
        print(f"错误：文件不存在 - {input_file}")
        return False

    print(f"正在读取文件: {input_file}")
    print("="*80)

    # 加载现有的Excel文件
    try:
        wb = load_workbook(input_file)
    except Exception as e:
        print(f"错误：无法读取Excel文件 - {str(e)}")
        return False

    # 检查是否存在"所有PDF文件"工作表
    if "所有PDF文件" not in wb.sheetnames:
        print("错误：未找到'所有PDF文件'工作表")
        return False

    ws_original = wb["所有PDF文件"]

    # 读取原始数据
    print("正在读取原始数据...")
    all_data = []
    headers = []

    # 读取标题行
    for cell in ws_original[1]:
        headers.append(cell.value)

    # 找到Email地址和Email数量的列索引
    email_col_idx = headers.index("Email地址") if "Email地址" in headers else -1
    email_count_col_idx = headers.index("Email数量") if "Email数量" in headers else -1

    if email_col_idx == -1:
        print("错误：未找到'Email地址'列")
        return False

    # 读取所有数据行
    for row in ws_original.iter_rows(min_row=2, values_only=True):
        all_data.append(list(row))

    print(f"读取到 {len(all_data)} 行数据")

    # 处理数据，过滤edu邮箱
    print("正在过滤包含'edu'的邮箱...")
    filtered_data = []
    total_original_emails = 0
    total_filtered_emails = 0
    files_with_changes = 0

    for row in all_data:
        new_row = list(row)
        original_email_str = row[email_col_idx]

        # 过滤邮箱
        filtered_emails = filter_edu_emails(original_email_str)

        # 统计
        if isinstance(original_email_str, str) and original_email_str != "未找到":
            original_count = len([e.strip() for e in original_email_str.split(';')])
            total_original_emails += original_count
        else:
            original_count = 0

        filtered_count = len(filtered_emails)
        total_filtered_emails += filtered_count

        if original_count != filtered_count:
            files_with_changes += 1

        # 更新邮箱列
        if filtered_emails:
            new_row[email_col_idx] = "; ".join(filtered_emails)
            new_row[email_count_col_idx] = filtered_count
        else:
            # 区分两种情况：原本就没有 vs 全部被过滤
            if original_count > 0:
                # 原本有email但全部被过滤
                new_row[email_col_idx] = "已全部过滤（均为edu）"
            else:
                # 原本就没有email
                new_row[email_col_idx] = "未找到"
            new_row[email_count_col_idx] = 0

        filtered_data.append(new_row)

    print(f"过滤完成:")
    print(f"  原始Email总数: {total_original_emails}")
    print(f"  过滤后Email总数: {total_filtered_emails}")
    print(f"  过滤掉的Email数: {total_original_emails - total_filtered_emails}")
    print(f"  受影响的文件数: {files_with_changes}")

    # 创建新工作表
    print("\n正在创建新工作表...")
    ws_filtered = wb.create_sheet(title="已过滤Email（无edu）")

    # 设置样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    skipped_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 写入标题行
    ws_filtered.append(headers)

    # 设置标题行样式
    for cell in ws_filtered[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 写入过滤后的数据
    for row_data in filtered_data:
        ws_filtered.append(row_data)

    # 设置数据行样式
    for row_idx, row_data in enumerate(filtered_data, start=2):
        for cell in ws_filtered[row_idx]:
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            # 如果是跳过的文件，设置黄色背景
            status_col_idx = headers.index("处理状态") if "处理状态" in headers else -1
            if status_col_idx != -1 and row_data[status_col_idx] == '跳过':
                cell.fill = skipped_fill

    # 调整列宽（与原工作表相同）
    ws_filtered.column_dimensions['A'].width = 8
    ws_filtered.column_dimensions['B'].width = 50
    ws_filtered.column_dimensions['C'].width = 25
    ws_filtered.column_dimensions['D'].width = 70
    ws_filtered.column_dimensions['E'].width = 70
    ws_filtered.column_dimensions['F'].width = 15
    ws_filtered.column_dimensions['G'].width = 12
    ws_filtered.column_dimensions['H'].width = 40
    ws_filtered.column_dimensions['I'].width = 12
    ws_filtered.column_dimensions['J'].width = 35

    # 更新或创建统计工作表
    print("正在更新统计信息...")
    if "统计汇总" in wb.sheetnames:
        ws_stats = wb["统计汇总"]
        # 在现有统计后添加过滤统计
        last_row = ws_stats.max_row
        ws_stats.append([])  # 空行
        ws_stats.append(["过滤后统计（无edu邮箱）", ""])
        ws_stats.append(["过滤后Email总数", total_filtered_emails])
        ws_stats.append(["过滤掉的Email数", total_original_emails - total_filtered_emails])
        ws_stats.append(["受影响的文件数", files_with_changes])
        ws_stats.append(["过滤时间", datetime.now().strftime('%Y-%m-%d %H:%M:%S')])

        # 设置样式
        for row in ws_stats.iter_rows(min_row=last_row+2, max_row=ws_stats.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical='center')

        # 设置标题样式
        for cell in ws_stats[last_row+2]:
            cell.fill = header_fill
            cell.font = header_font

    # 保存为新文件
    output_file = input_file.replace('.xlsx', '_已过滤.xlsx')
    if output_file == input_file:
        output_file = input_file.replace('.xlsx', '') + '_已过滤.xlsx'

    print(f"\n正在保存文件: {output_file}")
    try:
        wb.save(output_file)
        print("✓ 文件保存成功！")
        print("="*80)
        print(f"\n新文件已生成: {output_file}")
        print(f"\n文件包含 {len(wb.sheetnames)} 个工作表:")
        for idx, sheet_name in enumerate(wb.sheetnames, 1):
            print(f"  {idx}. {sheet_name}")
        return True
    except Exception as e:
        print(f"错误：无法保存文件 - {str(e)}")
        return False

def main():
    """主函数"""
    print("="*80)
    print("EDU邮箱过滤工具")
    print("="*80)
    print()

    # 确定输入文件
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
    else:
        # 默认处理当前目录下的文件
        input_file = "PDF_Email提取报告.xlsx"
        if not os.path.exists(input_file):
            print("使用方法:")
            print("  python filter_edu_emails.py [Excel文件路径]")
            print()
            print("示例:")
            print("  python filter_edu_emails.py")
            print("    -> 处理当前目录的 PDF_Email提取报告.xlsx")
            print()
            print("  python filter_edu_emails.py \"D:\\报告\\PDF_Email提取报告.xlsx\"")
            print("    -> 处理指定路径的文件")
            print()
            print(f"错误：当前目录未找到 {input_file}")
            return

    # 处理文件
    success = process_excel(input_file)

    if success:
        print("\n✓ 处理完成！")
        print("\n说明:")
        print("  - 原文件保持不变")
        print("  - 新文件添加了'已过滤Email（无edu）'工作表")
        print("  - 可以对比两个工作表查看过滤效果")
    else:
        print("\n✗ 处理失败，请检查错误信息")

if __name__ == "__main__":
    main()
